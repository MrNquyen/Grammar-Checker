# app.py
import openpyxl
import xlwings as xw
import pandas as pd
import numpy as np
import os
from flask import Flask, render_template, url_for, request, send_file, session, jsonify
from flask_sqlalchemy import SQLAlchemy
from icecream import ic
from datetime import timedelta
from io import BytesIO
from utils.general import load_yml_to_args, load_yml, clean_local_path 
from utils.excel_utils import load_excel_wb, onedrive_url_to_iframe, finalize
from utils.logger import Logger
from utils.configs import Config
from utils.registry import registry
from utils.session import switch_current
from utils.chatbot import get_correction_results
from projects.agent.agent_checker import GramCheckerAgent
from utils.history_handler import HistoryHandler

# Configuration
save_dir = "save/upload"

# Flask app
agent_checker = GramCheckerAgent()
history_handler = HistoryHandler()

app = Flask(__name__)
app.secret_key = "supersecret"  # required for session to work


@app.route("/")
def index():
    # session.clear()
    all_current_files = history_handler.get_all_current_files()
    return render_template("index.html", all_files=all_current_files)


#-- Upload files
@app.route("/upload", methods=["POST"])
def upload():
    #~ Get upload file
    local_path = request.form.get("local_path", "").strip()
    local_path = clean_local_path(local_path)
    excel_url = request.form.get("excel_url", "").strip()

    #~ Load and save
    error_message = None
    if not local_path:
        error_message = "Please enter a local path file"
    if not excel_url:
        error_message = "Please enter a OneDrive excel url file"
    if not os.path.isfile(local_path):
        error_message = "Please enter a valid local path"
    else:
        wb = load_excel_wb(local_path)
        #~ Save to session
        session["current_excel_file_path"] = local_path
        session["current_excel_url"] = excel_url
        session["sheet_names"] = wb.sheetnames
        session["current_sheet_name"] = session["sheet_names"][0]
        session["current_iframe"] = onedrive_url_to_iframe(
            url=excel_url,
            sheetname=session["current_sheet_name"]
        )

        #~ Save to database
        history_handler.add_file_information(
            local_path=local_path,
            online_url=excel_url,
            iframe=session["current_iframe"],
            sheetnames=session["sheet_names"]
        )
    all_current_files = history_handler.get_all_current_files()
    return render_template("index.html", error_message=error_message, all_files=all_current_files)


#-- Select Files
@app.route("/select_file", methods=["POST"])
def select_file():
    selected_local_path = request.args.get("selected_local_path")
    file_info = history_handler.get_file_information(selected_local_path)
    session["current_excel_file_path"] = file_info["local_path"]
    session["current_excel_url"] = file_info["online_url"]
    session["sheet_names"] = file_info["sheetnames"]
    session["current_sheet_name"] = file_info["sheetnames"][0]
    session["current_iframe"] = file_info["iframe"]


#-- Show Excel Files on Web
@app.route("/show_excel", methods=["GET", "POST"])
def show_excel():
    selected_file = request.form.get("selected_file")
    if not selected_file:
        error_message = "Please select the file first"
        return render_template("index.html", error_message=error_message)
    
    #~ Switch to selected file
    switch_current(history_handler, selected_file)
    
    #~ Render template
    return render_template("show_excel.html", sheets=session["sheet_names"], iframe=session["current_iframe"])



#-- Show sheet
@app.route("/show_sheet", methods=["GET", "POST"])
def show_sheet():
    current_sheet_name = request.args.get("sheet")
    if "current_excel_file_path" not in session:
        return "Please upload the file first"

    #-- Get iframe
    session["current_sheet_name"] = current_sheet_name
    session["current_iframe"] = onedrive_url_to_iframe(
        url=session["current_excel_url"],
        sheetname=session["current_sheet_name"]
    )

    #-- Get error-correction-list
    correction_results = history_handler.get_correction_history_info(
        local_path=session["current_excel_file_path"],
        sheet_name=session["current_sheet_name"]
    )
    ic(correction_results)
    #-- Return
    return render_template(
        "show_excel.html", 
        sheets=session["sheet_names"], 
        iframe=session["current_iframe"],
        current_sheet_name=session["current_sheet_name"],
        correction_results=correction_results
    )

#-- Error List
@app.route("/show_sheet_cell", methods=["POST"])
def show_sheet_cell():
    data = request.get_json()
    cell = data.get("cell")
    if "current_excel_file_path" not in session:
        return "Please upload the file first"

    session["current_iframe"] = onedrive_url_to_iframe(
        url=session["current_excel_url"],
        sheetname=session["current_sheet_name"],
        current_cell=cell
    )
    ic(session["current_iframe"])

    return jsonify({
        "iframe": session.get("current_iframe", ""),
        "current_sheet_name": session.get("current_sheet_name", "")
    })


@app.route("/change_sheet_cell", methods=["POST"])
def change_sheet_cell():
    data = request.get_json()
    cell = data.get("cell")
    old_value = data.get("old_value")
    new_value = data.get("new_value")

    if old_value != new_value:
        #~ Load wb and move get current_sheet
        wb = xw.Book(session["current_excel_file_path"])
        current_sheet = wb.sheets[session["current_sheet_name"]]

        #~ Change the cell value
        agent_checker.change_sheet_cell(
            sheet=current_sheet,
            cell_address=cell, 
            old_value=old_value, 
            new_value=new_value
        )
        finalize(wb)
    else:
        ic("Old value is similar with new value")

    return jsonify({
        "iframe": session.get("current_iframe", ""),
        "current_sheet_name": session.get("current_sheet_name", "")
    })

    

#---- Serve for SpreadJS
@app.route("/download_excel")
def download_excel():
    return send_file(session["current_excel_file_path"], as_attachment=True)


@app.route("/check_grammar", methods=["POST", "GET"])
def check_grammar():
    # Need setup database 
    ic("Start Checking grammar")
    data = request.get_json()  # Expect JSON payload
    request_sheet_name = data["sheet_name"]

    # Load excel
    wb = load_excel_wb(path=session["current_excel_file_path"])

    # Check sheet grammar
    rows = [row for row in wb[request_sheet_name].iter_rows(values_only=True)]
    checked_rows = agent_checker.run_sheet(rows)
    
    results = get_correction_results(
        old_rows=rows,
        new_rows=checked_rows,
    )

    # Add correction history to database
    file_id = history_handler.get_file_id(session["current_excel_file_path"])
    ic("Add to history")
    for result in results:
        history_handler.add_correction_history(
            file_id=file_id, 
            sheet_name=session["current_sheet_name"], 
            coordinates=result["coordinates"], 
            old_value=result["old_value"], 
            new_value=result["new_value"]
        )

    ic(results)
    ic("Finish Check Grammar")
    return jsonify({
        "iframe": session.get("current_iframe", ""),
        "results":results
    }) 


# ==== BASE SETTING ====
class SystemBase:
    def __init__(self, config_run_path):
        #~ Configuration
        args = load_yml_to_args(config_run_path)
        self.args = args
        #~ Build
        self.build()

    #-- BUILD
    def build(self):
        self.build_logger()
        self.build_config()
        self.build_registry()

    def build_config(self):
        self.config = Config(load_yml(self.args.config))
        
    def build_logger(self):
        self.writer = Logger(name="general")

    def build_registry(self):
        #~ Build writer
        registry.set_module("writer", name="common", instance=self.writer)
        #~ Build args
        registry.set_module("args", name=None, instance=self.args)
        #~ Build config
        self.config.build_registry()



if __name__=="__main__":
    config_run_path = r"config/config_run.yaml"
    system = SystemBase(config_run_path)
    app.run(debug=True)
