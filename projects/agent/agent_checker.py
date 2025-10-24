from projects.agent.agent_base import BaseAgent
from langchain.chains import LLMChain
from tqdm import tqdm
from typing import List
from icecream import ic
import openpyxl
import json
import asyncio
import nest_asyncio
import re
import numpy as np
from langchain_core.prompts import ChatPromptTemplate, PromptTemplate
from langchain_core.output_parsers import JsonOutputParser
from langchain_core.runnables import RunnableSequence, RunnableLambda



nest_asyncio.apply()
semaphore = asyncio.Semaphore(8) 

class GramCheckerAgent(BaseAgent):
    ROW_GRAMMAR_CHECK_SYSTEM = """
        You are an English proofreader and editor.
        Your task is to correct spelling, grammar, and sentence fluency in the given list of texts.
        You may improve clarity and naturalness, but you must NOT remove or omit any part of the original content.
    """

    ROW_GRAMMAR_CHECK_INSTRUCTION = """
        ### Instructions:
        You are given a list of texts. Your task is to check and improve each text based on grammar, spelling, and readability, while ensuring that NO content or elements from the original text are deleted or removed.

        ### Rules:
        1. For each text in the input list:
            - Correct only **grammar**, **spelling**, and incorrect **word usage**.
            - Keep all spacing, line breaks ("\\n"), and special symbols as they appear in the original.
            
        2. Output as a **structured JSON format** that JsonOutputParser can parsed.

        ### Notes:
            - Do NOT add or remove any content, words, or symbols.
            - Do NOT correct the error related to symbols or whitespace.

        ### Output format:
        {{
            "data": [
                {{
                    "text_id" (int): <ID of the input text>,
                    "status" (bool): <False if any issues were found and fixed, True if no errors>,
                    "fixed_text" (str): <corrected and improved text>,
                    "original_text" (str): <original input text>
                }},
                ...
            ]
        }}

        ### Input:
        {input_list_text}
    """


    def __init__(self):
        super().__init__()
        self.create_fixed_chain()


    #-- Grammar Checker
    def create_fixed_chain(self):
        ROW_GRAMMAR_CHECK_PROMPT = self.ROW_GRAMMAR_CHECK_SYSTEM + self.ROW_GRAMMAR_CHECK_INSTRUCTION
        ROW_GRAMMAR_CHECK_PROMPT_TEMPLATE = PromptTemplate(
            template=ROW_GRAMMAR_CHECK_PROMPT,
            input_variables=["input_list_text"]
        )

        sanitize_step = RunnableLambda(lambda x: self.sanitize_json(x))
        parser = JsonOutputParser()
        self.chain = (
            ROW_GRAMMAR_CHECK_PROMPT_TEMPLATE
            | self.gpt_llm
            | parser       
        )
        # self.chain = ROW_GRAMMAR_CHECK_PROMPT_TEMPLATE | self.gpt_llm | parser


    async def check_list(self, list_texts):
        def format_input_to_json(list_texts):
            rewrite_list_texts = []
            for id, text in enumerate(list_texts):
                rewrite_list_texts.append({
                    "id": id + 1,
                    "text": text
                })
            return rewrite_list_texts

        rewrite_list_texts = format_input_to_json(list_texts)
        response = await self.chain.ainvoke({"input_list_text": rewrite_list_texts})

        if "text" in response:
            response = response["text"]
        if "data" in response:
            response = response["data"]
        return response


    async def corrected_sheet(self, rows, batch_size=15):
        if not isinstance(rows, np.ndarray):
            rows = np.array(rows, dtype=object)

        org_shape = rows.shape
        flatten_rows = rows.flatten()
        non_nan_indices  = [i for i, v in enumerate(flatten_rows) if v is not None]
        non_nan_rows = flatten_rows[non_nan_indices]

        async def limited_check(list_texts):
            async with semaphore:
                await asyncio.sleep(0.4)
                return await self.check_list(list_texts)

        #-- Batch iteration
        tasks = []
        all_indices = []
        for start_idx in range(0, len(non_nan_rows), batch_size):
            batch_non_nan_rows = non_nan_rows[start_idx: start_idx + batch_size]
            batch_non_nan_indices = non_nan_indices [start_idx: start_idx + batch_size]
            all_indices.extend(batch_non_nan_indices)
            tasks.append(limited_check(list_texts=batch_non_nan_rows))
            
        all_batch_responses = await asyncio.gather(*tasks)
        all_responses = [response for batch_responses in all_batch_responses for response in batch_responses]

        for idx, response in zip(all_indices, all_responses):
            if not response["status"]:
                if "text" in response:
                    response = response["text"]
                if "data" in response:
                    response = response["data"]
                if not str(response["fixed_text"]).strip() == str(response["original_text"]).strip():
                    flatten_rows[idx] = response["fixed_text"]
        new_rows = flatten_rows.reshape(org_shape)
        return new_rows



    async def check_all_sheet(self, wb):
        sheet_names = wb.sheetnames
        tasks = []
        for sheet_name in sheet_names:
            sheet = wb[sheet_name]
            rows = [np.array(row) for row in sheet.iter_rows(values_only=True)]
            tasks.append(corrected_sheet(rows))
        new_sheets = await asyncio.gather(*tasks)
        return new_sheets

    
    def run_all_sheet(self, excel_path: str, is_saved):
        if not excel_path.endswith(".xlsx"):
            raise ValueError("Invalid files")
        wb = openpyxl.load_workbook(filename=excel_path)
        sheet_names = wb.sheetnames

        all_new_sheets = asyncio.run(self.check_all_sheet(wb))
        return all_new_sheets


    def run_sheet(self, sheet_rows: List):
        new_sheet_rows = asyncio.run(self.corrected_sheet(sheet_rows))
        return new_sheet_rows


    #-- Interact with cell and sheet
    def find_all_common_substring(self, string_a, string_b):
        list_a = string_a.split(" ")
        list_b = string_b.split(" ")
        n, m = len(list_a), len(list_b)

        #~ DP Matrix
        matrix = np.zeros((n+1, m+1), dtype=int)
        for i in range(n):
            for j in range(m):
                if list_a[i] == list_b[j]:
                    matrix[i+1, j+1] = matrix[i, j] + 1

        #~ Collect substrings with indices
        candidates = []
        for i in range(1, n+1):
            for j in range(1, m+1):
                length = matrix[i, j]
                if length > 0:
                    substring = " ".join(list_a[i-length:i])
                    start_a = i - length
                    start_b = j - length
                    candidates.append((substring, start_a, i, start_b, j))  

        #~ Keep only maximal ones
        maximal = []
        for s, sa, ea, sb, eb in candidates:
            if not any((s != t and s in t) for t, *_ in candidates):
                maximal.append((s, sa, ea, sb, eb))

        #~ Deduplicate
        seen = set()
        result = []
        for item in maximal:
            if item[0] not in seen:
                result.append(item)
                seen.add(item[0])
        return result

    
    def get_common_missing_idx(self, string_a, string_b):
        result = self.find_all_common_substring(string_a=string_a, string_b=string_b)
        #-- Get missing idx
        len_a = len(string_a.split(" "))
        len_b = len(string_b.split(" "))
        string_a_idx_mask = np.zeros(len_a, dtype=bool)
        string_b_idx_mask = np.zeros(len_b, dtype=bool)

        for _, sa, ea, sb, eb in result:
            string_a_idx_mask[sa:ea] = True
            string_b_idx_mask[sb:eb] = True

        common_a = np.flatnonzero(string_a_idx_mask)
        common_b = np.flatnonzero(string_b_idx_mask)
        missing_a = np.flatnonzero(~string_a_idx_mask)
        missing_b = np.flatnonzero(~string_b_idx_mask)

        return common_a, common_b, missing_a, missing_b, result


    def get_character_styles(self, sheet, cell_address):
        cell = sheet.range(cell_address)
        text = cell.value
        if not text:
            return styles
        
        #~ We'll group characters by their formatting runs:
        styles = []
        length = len(text)
        words = text.split(" ")
        word_info = {}
        i = 0
        for word_idx, word in enumerate(words):
            len_word = len(word)
            list_char_idx = list(range(i, i + len_word + 1))
            i += len_word + 1
            word_info[word_idx] = {
                "list_char_idx": list_char_idx,
                "styles": None,
            }

        for word_idx, word_item_info in word_info.items():
            start_idx = min(word_info[word_idx]["list_char_idx"] )
            first_char_font = cell.characters[start_idx].font
            word_info[word_idx]["styles"] = {
                "bold": first_char_font.bold,
                "italic": first_char_font.italic,
                "color": first_char_font.color,
                "size": first_char_font.size,
                "name": first_char_font.name,
            }
        return word_info


    # def set_styles(self, modify_sheet, cell_address, root_word_info, common_root, common_modify):
    #     modify_word_info = self.get_character_styles(modify_sheet, cell_address)
    #     modify_cell = modify_sheet.range(cell_address)
    #     for root_idx, modify_idx in tqdm(zip(common_root, common_modify)):
    #         word_at_root_idx_info = root_word_info[root_idx]
    #         word_at_modify_idx_info = modify_word_info[modify_idx]

    #         modify_styles = word_at_root_idx_info["styles"]
    #         modify_list_char_idx = word_at_modify_idx_info["list_char_idx"] 
    #         for modify_char_idx in modify_list_char_idx:
    #             modify_cell_font = modify_cell.characters[modify_char_idx].font
    #             for attr, value in modify_styles.items():
    #                 setattr(modify_cell_font, attr, value)

    def set_styles(self, modify_sheet, cell_address, common_root, common_modify, new_value):
        root_word_info = self.get_character_styles(modify_sheet, cell_address)
        modify_cell = modify_sheet.range(cell_address)
        modify_cell.value = new_value

        modify_word_info = self.get_character_styles(modify_sheet, cell_address)
        for root_idx, modify_idx in tqdm(zip(common_root, common_modify)):
            word_at_root_idx_info = root_word_info[root_idx]
            word_at_modify_idx_info = modify_word_info[modify_idx]

            modify_styles = word_at_root_idx_info["styles"]
            modify_list_char_idx = word_at_modify_idx_info["list_char_idx"] 
            for modify_char_idx in modify_list_char_idx:
                modify_cell_font = modify_cell.characters[modify_char_idx].font
                for attr, value in modify_styles.items():
                    setattr(modify_cell_font, attr, value)


    def change_sheet_cell(self, sheet, cell_address, old_value, new_value):
        common_old, common_new, missing_old, missing_new, result = self.get_common_missing_idx(old_value, new_value)
        self.set_styles(
            modify_sheet=sheet,
            cell_address=cell_address,
            common_root=common_old,
            common_modify=common_new,
            new_value=new_value
        )

    #-- PREPROCESSING
    def sanitize_json(self, raw_output: str):
        if isinstance(raw_output, dict):
            return json.dumps(raw_output, ensure_ascii=False)

        raw = str(raw_output).strip()
        raw = re.sub(r"^```(?:json)?", "", raw)
        raw = re.sub(r"```$", "", raw)
        raw = raw.strip()
        raw = re.sub(r'^[^{\[]*', '', raw)
        raw = re.sub(r'[^}\]]*$', '', raw)

        try:
            parsed = json.loads(raw)
            compact = json.dumps(parsed, ensure_ascii=False, indent=2)
            return compact
        except Exception:
            return raw




